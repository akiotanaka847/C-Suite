from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Any

from openexecutive.knowledge.store import ChromaDBStore

logger = logging.getLogger(__name__)

BUILTIN_KNOWLEDGE_PATH = Path(__file__).parent / "builtin"
FAILURES_KNOWLEDGE_PATH = BUILTIN_KNOWLEDGE_PATH / "failures"

DOMAIN_MAP: dict[str, str] = {
    "strategy": "strategy",
    "finance": "finance",
    "hr": "hr",
    "legal": "legal",
    "operations": "operations",
    "marketing": "marketing",
    "board": "board",
    "product": "product",
}


def chunk_text(text: str, chunk_size: int = 512, overlap: int = 50) -> list[str]:
    words = text.split()
    if not words:
        return []

    chunks: list[str] = []
    start = 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        chunk = " ".join(words[start:end])
        chunks.append(chunk)
        if end == len(words):
            break
        start = end - overlap
    return chunks


def extract_text_from_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n\n".join(pages)


def extract_text_from_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_text_from_xlsx(path: Path, max_chars: int = 200_000) -> str:
    """Flatten an .xlsx/.xlsm workbook to text — one ``## <sheet>`` heading per
    NON-EMPTY worksheet, cells tab-joined and rows newline-joined. Only reads
    stored cell values (``data_only=True`` returns cached formula results, not
    formulae); legacy binary ``.xls`` is not supported by openpyxl.

    ``read_only`` streams rows and ``max_chars`` bounds the accumulated text, so
    a decompression-bombed workbook (a small archive that inflates to millions
    of cells) can't exhaust memory."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        total = 0
        for ws in wb.worksheets:
            heading_written = False
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if not cells:
                    continue
                if not heading_written:
                    # Defer the heading until the sheet is known to have data,
                    # so a fully-blank sheet contributes nothing.
                    heading = f"## {ws.title}"
                    parts.append(heading)
                    total += len(heading) + 1
                    heading_written = True
                line = "\t".join(cells)
                parts.append(line)
                total += len(line) + 1
                if total >= max_chars:
                    return "\n".join(parts)
        return "\n".join(parts)
    finally:
        wb.close()


def extract_text_from_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_text_from_pdf(path)
    elif suffix in (".docx", ".doc"):
        return extract_text_from_docx(path)
    elif suffix in (".xlsx", ".xlsm"):
        return extract_text_from_xlsx(path)
    elif suffix in (".md", ".txt", ".rst", ".csv"):
        return path.read_text(encoding="utf-8")
    return ""


def _make_chunk_id(source: str, chunk_index: int) -> str:
    base = f"{source}::chunk::{chunk_index}"
    return hashlib.md5(base.encode()).hexdigest()


def infer_domain_from_path(path: Path) -> str:
    for part in path.parts:
        domain = DOMAIN_MAP.get(part.lower())
        if domain:
            return domain
    return "general"


async def ingest_file(
    path: Path,
    store: ChromaDBStore,
    domain: str | None = None,
    collection: str = ChromaDBStore.COMPANY_COLLECTION,
) -> int:
    text = extract_text_from_file(path)
    if not text.strip():
        return 0

    inferred_domain = domain or infer_domain_from_path(path)
    chunks = chunk_text(text, chunk_size=512, overlap=50)

    texts = chunks
    metadatas: list[dict[str, Any]] = [
        {
            "domain": inferred_domain,
            "filename": path.name,
            "source": str(path),
            "chunk_index": i,
        }
        for i in range(len(chunks))
    ]
    ids = [_make_chunk_id(str(path), i) for i in range(len(chunks))]

    store.add_documents(texts=texts, metadatas=metadatas, ids=ids, collection=collection)
    return len(chunks)


async def ingest_text(
    text: str,
    store: ChromaDBStore,
    *,
    source_name: str,
    domain: str = "general",
    collection: str = ChromaDBStore.COMPANY_COLLECTION,
    extra_metadata: dict[str, Any] | None = None,
) -> int:
    """Ingest a raw markdown/text string as knowledge (no file on disk).

    Mirrors ``ingest_file`` but takes a string — used to persist the
    executive_research artifact into its own collection. ``source_name``
    is the logical identifier used for both the ``filename``/``source``
    metadata and the chunk-id namespace. ``extra_metadata`` is merged into
    every chunk's metadata (e.g. ``{"type": "recent_research", "created_at": …}``).
    Returns the number of chunks written.
    """
    if not text.strip():
        return 0

    chunks = chunk_text(text, chunk_size=512, overlap=50)
    extra = extra_metadata or {}
    metadatas: list[dict[str, Any]] = [
        {
            "domain": domain,
            "filename": source_name,
            "source": source_name,
            "chunk_index": i,
            **extra,
        }
        for i in range(len(chunks))
    ]
    ids = [_make_chunk_id(source_name, i) for i in range(len(chunks))]

    store.add_documents(texts=chunks, metadatas=metadatas, ids=ids, collection=collection)
    return len(chunks)


async def ingest_builtin_file(
    path: Path,
    store: ChromaDBStore,
    collection: str = ChromaDBStore.BUILTIN_COLLECTION,
    chunk_type: str = "builtin",
    chunk_size: int = 512,
    overlap: int = 50,
) -> int:
    """Index a single built-in markdown file. Caller must delete old chunks first.

    Defaults match the positive-playbook ingest path. Pass
    ``collection=ChromaDBStore.FAILURES_COLLECTION`` (with ``chunk_type='failure_case'``
    and smaller chunks) to ingest a single failure case study — keeps the
    failure CRUD endpoints in lockstep with ``seed_failures``.
    """
    text = path.read_text(encoding="utf-8")
    if not text.strip():
        return 0
    domain = infer_domain_from_path(path)
    chunks = chunk_text(text, chunk_size=chunk_size, overlap=overlap)
    metadatas: list[dict[str, Any]] = [
        {
            "domain": domain,
            "filename": path.name,
            "source": str(path),
            "chunk_index": i,
            "type": chunk_type,
        }
        for i in range(len(chunks))
    ]
    ids = [_make_chunk_id(str(path), i) for i in range(len(chunks))]
    store.add_documents(
        texts=chunks,
        metadatas=metadatas,
        ids=ids,
        collection=collection,
    )
    return len(chunks)


def list_company_docs(docs_dir: Path) -> list[dict[str, Any]]:
    if not docs_dir.exists():
        return []
    return [
        {
            "filename": f.name,
            "size_bytes": f.stat().st_size,
            "modified_at": f.stat().st_mtime,
        }
        for f in sorted(docs_dir.iterdir())
        if f.is_file() and not f.name.startswith(".")
    ]


async def seed_builtin_knowledge(
    store: ChromaDBStore | None = None,
    force: bool = False,
) -> int:
    if store is None:
        from openexecutive.config import get_settings

        settings = get_settings()
        store = ChromaDBStore(persist_directory=settings.vector_store_path)

    if not force and store.get_collection_count(ChromaDBStore.BUILTIN_COLLECTION) > 0:
        return 0

    total = 0
    for md_file in BUILTIN_KNOWLEDGE_PATH.rglob("*.md"):
        # Skills live under builtin/skills/ but are indexed into a separate
        # collection by openexecutive.knowledge.skills_index.seed_builtin_skills.
        if any(p in md_file.relative_to(BUILTIN_KNOWLEDGE_PATH).parts for p in ("skills", "failures")):
            continue
        domain = infer_domain_from_path(md_file)
        text = md_file.read_text(encoding="utf-8")
        chunks = chunk_text(text)

        metadatas: list[dict[str, Any]] = [
            {
                "domain": domain,
                "filename": md_file.name,
                "source": str(md_file),
                "chunk_index": i,
                "type": "builtin",
            }
            for i in range(len(chunks))
        ]
        ids = [_make_chunk_id(str(md_file), i) for i in range(len(chunks))]
        store.add_documents(
            texts=chunks,
            metadatas=metadatas,
            ids=ids,
            collection=ChromaDBStore.BUILTIN_COLLECTION,
        )
        total += len(chunks)

    return total


async def seed_failures(
    store: ChromaDBStore | None = None,
    force: bool = False,
) -> int:
    """Index all failure case studies from builtin/failures/<domain>/*.md.

    Idempotent: skipped if the failures collection is already non-empty,
    unless force=True. Uses a smaller chunk size (400 words) to preserve
    the narrative arc of each section (situation/root-cause/lessons).
    """
    if store is None:
        from openexecutive.config import get_settings

        settings = get_settings()
        store = ChromaDBStore(persist_directory=settings.vector_store_path)

    if not force and store.get_collection_count(ChromaDBStore.FAILURES_COLLECTION) > 0:
        return 0

    if not FAILURES_KNOWLEDGE_PATH.is_dir():
        logger.warning("failures knowledge path not found, skipping: %s", FAILURES_KNOWLEDGE_PATH)
        return 0

    total = 0
    for md_file in FAILURES_KNOWLEDGE_PATH.rglob("*.md"):
        domain = infer_domain_from_path(md_file)
        text = md_file.read_text(encoding="utf-8")
        if not text.strip():
            continue
        chunks = chunk_text(text, chunk_size=400, overlap=40)
        metadatas: list[dict[str, Any]] = [
            {
                "domain": domain,
                "filename": md_file.name,
                "source": str(md_file),
                "chunk_index": i,
                "type": "failure_case",
            }
            for i in range(len(chunks))
        ]
        ids = [_make_chunk_id(str(md_file), i) for i in range(len(chunks))]
        store.add_documents(
            texts=chunks,
            metadatas=metadatas,
            ids=ids,
            collection=ChromaDBStore.FAILURES_COLLECTION,
        )
        total += len(chunks)

    return total
