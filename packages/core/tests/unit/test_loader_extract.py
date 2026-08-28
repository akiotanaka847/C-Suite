"""Text extraction for company-doc / intake ingestion.

Covers the formats added for client-intake attachments (xlsx + csv) and the
``extract_text_from_file`` dispatcher's behavior on an unsupported suffix.
"""
from __future__ import annotations

from pathlib import Path

from openexecutive.knowledge.loader import (
    extract_text_from_file,
    extract_text_from_xlsx,
)


def test_extract_text_from_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws.append(["Metric", "Value"])
    ws.append([None, None])  # all-empty row is skipped
    ws.append(["ARR", 1200000])
    # A fully-blank second sheet must contribute no heading.
    wb.create_sheet("Empty")
    path = tmp_path / "sheet.xlsx"
    wb.save(str(path))

    text = extract_text_from_xlsx(path)
    assert "## Financials" in text
    assert "Metric\tValue" in text
    assert "ARR\t1200000" in text
    # Blank rows and blank sheets contribute nothing.
    assert "## Empty" not in text
    assert "\n\n" not in text

    # Routed through the dispatcher identically.
    assert extract_text_from_file(path) == text


def test_extract_text_from_xlsx_respects_char_cap(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for i in range(100):
        ws.append([f"row{i}", "x" * 50])
    path = tmp_path / "big.xlsx"
    wb.save(str(path))

    capped = extract_text_from_xlsx(path, max_chars=200)
    assert len(capped) < 1000  # stopped well before flattening all 100 rows


def test_extract_text_from_csv(tmp_path: Path) -> None:
    path = tmp_path / "roster.csv"
    path.write_text("name,role\nDana Reyes,CEO\n", encoding="utf-8")
    assert "Dana Reyes,CEO" in extract_text_from_file(path)


def test_extract_text_unsupported_suffix(tmp_path: Path) -> None:
    path = tmp_path / "thing.bin"
    path.write_bytes(b"\x00\x01\x02")
    assert extract_text_from_file(path) == ""
